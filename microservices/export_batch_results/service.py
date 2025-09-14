
import os
import json
import json as _json
import pandas as pd
from flask import request, jsonify
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import BarChart, Reference, PieChart, Series
from tempfile import NamedTemporaryFile
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
import io
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import pickle

def export_batch_results(request):
    data = request.get_json()
    batch_id = data.get('batch_id') if data else None
    if not batch_id:
        return jsonify({
            "status_code": 400,
            "status_message": "BAD REQUEST",
            "message": "batch_id is required in JSON body."
        }), 400
    # Find batch info in batch_data.json
    BATCH_DATA_PATH = os.path.abspath(os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'db_jsons', 'batch_data.json'))
    if not os.path.exists(BATCH_DATA_PATH):
        return jsonify({
            "status_code": 404,
            "status_message": "NOT FOUND",
            "message": "No batch data found."
        }), 404
    with open(BATCH_DATA_PATH, 'r', encoding='utf-8') as f:
        batch_data = json.load(f)
    batch_info = next((b for b in batch_data if b.get('batch_id') == batch_id), None)
    if not batch_info:
        return jsonify({
            "status_code": 404,
            "status_message": "NOT FOUND",
            "message": f"Batch with id '{batch_id}' not found."
        }), 404
    # Read result files
    def read_json_file(path):
        if not os.path.exists(path):
            return []
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    matched = read_json_file(batch_info['matched_data']['file_path'])
    suspected = read_json_file(batch_info['suspected_data']['file_path'])
    unmatched_source = read_json_file(batch_info['unmatched_source_data']['file_path'])
    unmatched_target = read_json_file(batch_info['unmatched_target_data']['file_path'])
    # Prepare DataFrames
    df_matched = pd.DataFrame(matched)
    # Export suspected groups as one row per group, with targets as JSON string
    suspected_for_excel = []
    for group in suspected:
        suspected_for_excel.append({
            'source_index': group.get('source_index'),
            'source_record': _json.dumps(group.get('source_record'), ensure_ascii=False),
            'targets': _json.dumps(group.get('targets'), ensure_ascii=False)
        })
    df_suspected = pd.DataFrame(suspected_for_excel)
    df_unmatched_source = pd.DataFrame(unmatched_source)
    df_unmatched_target = pd.DataFrame(unmatched_target)
    # Prepare report metrics
    total_source = len(df_matched) + len(df_suspected) + len(df_unmatched_source)
    total_target = len(df_matched) + len(df_suspected) + len(df_unmatched_target)
    report_data = [
        ["Batch ID", batch_id],
        ["Batch Name", batch_info.get('batch_name', '')],
        ["Total Source Records", total_source],
        ["Total Target Records", total_target],
        ["Matched Records", len(df_matched)],
        ["Suspected Records", len(df_suspected)],
        ["Unmatched Source Records", len(df_unmatched_source)],
        ["Unmatched Target Records", len(df_unmatched_target)],
        ["Match Rate (%)", round(100 * len(df_matched) / total_source, 2) if total_source else 0],
        ["Suspected Rate (%)", round(100 * len(df_suspected) / total_source, 2) if total_source else 0],
        ["Unmatched Source Rate (%)", round(100 * len(df_unmatched_source) / total_source, 2) if total_source else 0],
        ["Unmatched Target Rate (%)", round(100 * len(df_unmatched_target) / total_target, 2) if total_target else 0],
    ]
    # Create Excel workbook
    wb = Workbook()
    # Report sheet
    ws_report = wb.active
    ws_report.title = "Report"
    # Header formatting
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")  # Dark blue
    header_font = Font(bold=True, color="FFFFFF")  # White
    alt_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")  # Lighter blue
    for i, (k, v) in enumerate(report_data, 1):
        cell_key = ws_report.cell(row=i, column=1, value=k)
        cell_val = ws_report.cell(row=i, column=2, value=v)
        cell_key.fill = header_fill if i % 2 == 1 else alt_fill
        cell_val.fill = header_fill if i % 2 == 1 else alt_fill
        cell_key.font = header_font
        cell_val.font = header_font
        cell_key.alignment = Alignment(horizontal="left", vertical="center")
        cell_val.alignment = Alignment(horizontal="left", vertical="center")
    ws_report.column_dimensions['A'].width = 28
    ws_report.column_dimensions['B'].width = 20
    # Helper to add a styled sheet
    def add_sheet(name, df, suspected_raw=None):
        ws = wb.create_sheet(title=name)
        if name == "Suspected" and suspected_raw is not None:
            # Custom header
            ws.append(["source_index", "source_record", "targets"])
            for group in suspected_raw:
                ws.append([
                    group.get('source_index'),
                    _json.dumps(group.get('source_record'), ensure_ascii=False),
                    _json.dumps(group.get('targets'), ensure_ascii=False)
                ])
            ws.auto_filter.ref = ws.dimensions
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max(12, min(max_length + 2, 40))
            return
        if df.empty:
            ws.append(["No data available"])
            return
        # Header
        for col_idx, col in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col)
            cell.fill = header_fill if col_idx % 2 == 1 else alt_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # Data
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                if isinstance(value, dict):
                    value = _json.dumps(value, ensure_ascii=False)
                ws.cell(row=row_idx, column=col_idx, value=value)
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(12, min(max_length + 2, 40))
    add_sheet("Matched", df_matched)
    add_sheet("Suspected", df_suspected, suspected_raw=suspected)
    add_sheet("Unmatched_Source", df_unmatched_source)
    add_sheet("Unmatched_Target", df_unmatched_target)
    # Only show the match distribution pie chart, moved 5 rows below the Batch Summary
    ws_report = wb["Report"]
    pie_chart = PieChart()
    pie_chart.title = "Match Distribution"
    # The summary table starts at row 3, so chart should be at D8 (row 8, 5 rows below title)
    pie_data = Reference(ws_report, min_col=2, min_row=7, max_row=10)
    pie_labels = Reference(ws_report, min_col=1, min_row=7, max_row=10)
    pie_chart.add_data(pie_data, titles_from_data=False)
    pie_chart.set_categories(pie_labels)
    ws_report.add_chart(pie_chart, "D6")

    # Insert a blank row at the top, then Batch Summary title in row 2
    ws_report.insert_rows(1)
    ws_report.insert_rows(1)
    ws_report.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    title_cell = ws_report.cell(row=2, column=1, value="Batch Summary")
    title_cell.font = Font(bold=True, color="FFFFFF", size=13)
    title_cell.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Format keys as dark blue, bolded white font; values as black, not bold, starting from row 3
    for i, (k, v) in enumerate(report_data, 3):
        cell_key = ws_report.cell(row=i, column=1)
        cell_val = ws_report.cell(row=i, column=2)
        cell_key.font = Font(bold=True, color="FFFFFF")
        cell_key.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        cell_val.font = Font(bold=False, color="000000")
        cell_val.fill = PatternFill(fill_type=None)

    # --- Rule Insights Section ---
    # Gather rule info from matched and suspected
    rule_rows = []
    for rec in matched + suspected:
        rule_rows.append({
            'rule_id': rec.get('rule_id'),
            'rationale_statement': rec.get('rationale_statement')
        })
    import collections
    rule_ids = [r['rule_id'] for r in rule_rows if r['rule_id'] is not None]
    rule_counter = collections.Counter(rule_ids)
    most_common_rule = rule_counter.most_common(1)[0] if rule_counter else (None, 0)

    # Start Rule Insights section after the summary table
    rule_start_row = len(report_data) + 5
    ws_report.merge_cells(start_row=rule_start_row, start_column=1, end_row=rule_start_row, end_column=2)
    rule_title = ws_report.cell(row=rule_start_row, column=1, value="Rule Insights")
    rule_title.font = Font(bold=True, color="FFFFFF", size=13)
    rule_title.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    rule_title.alignment = Alignment(horizontal="center", vertical="center")

    # Add rule metrics
    metrics = [
        ("Unique Rules Applied", len(set(rule_ids))),
        ("Most Frequent Rule", str(most_common_rule[0]) if most_common_rule[0] else "-"),
        ("Most Frequent Rule Count", most_common_rule[1]),
    ]
    for j, (k, v) in enumerate(metrics, rule_start_row+1):
        cell_key = ws_report.cell(row=j, column=1, value=k)
        cell_val = ws_report.cell(row=j, column=2, value=v)
        cell_key.font = Font(bold=True, color="FFFFFF")
        cell_key.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        cell_val.font = Font(bold=False, color="000000")
        cell_val.fill = PatternFill(fill_type=None)

    # Save to temp file and upload to Google Drive
    # Save workbook to bytes buffer instead of file
    output = io.BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)

    # Google Drive upload
    SCOPES = ['https://www.googleapis.com/auth/drive.file']
    CREDENTIALS_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'credentials.json')
    TOKEN_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'token.pickle')
    creds = None
    # Load token if it exists
    if os.path.exists(TOKEN_PATH):
        with open(TOKEN_PATH, 'rb') as token:
            creds = pickle.load(token)
    # If no valid creds, do OAuth flow
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_PATH, SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for next run
        with open(TOKEN_PATH, 'wb') as token:
            pickle.dump(creds, token)
    drive_service = build('drive', 'v3', credentials=creds)
    # Find or create the 'Cross source linker - Exports' folder by name
    folder_name = 'Cross source linker - Exports'
    folder_id = None
    # Search for folder
    results = drive_service.files().list(q=f"mimeType='application/vnd.google-apps.folder' and name='{folder_name}' and trashed=false", fields="files(id, name)").execute()
    folders = results.get('files', [])
    if folders:
        folder_id = folders[0]['id']
    else:
        # Create folder if not found
        file_metadata_folder = {
            'name': folder_name,
            'mimeType': 'application/vnd.google-apps.folder'
        }
        folder = drive_service.files().create(body=file_metadata_folder, fields='id').execute()
        folder_id = folder.get('id')
    # Upload file to the folder using MediaIoBaseUpload to avoid file lock issues
    file_metadata = {
        'name': f"batch_{batch_id}_results.xlsx",
        'parents': [folder_id]
    }
    media = MediaIoBaseUpload(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    uploaded = drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()
    file_id = uploaded.get('id')

    # Make file shareable (anyone with link can view)
    drive_service.permissions().create(
        fileId=file_id,
        body={"role": "reader", "type": "anyone"},
    ).execute()
    file_link = f"https://drive.google.com/file/d/{file_id}/view?usp=sharing"

    return jsonify({
        "status_code": 200,
        "status_message": "SUCCESS",
        "message": f"Exported results for batch {batch_id}.",
        "file_name": f"batch_{batch_id}_results.xlsx",
        "file_link": file_link
    }), 200
